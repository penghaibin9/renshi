"""Real XLSX -> preview -> row commit -> HR02 occupancy, in the existing school lane.

Only permissions are seeded. The workbook is downloaded from the application;
all personnel facts must originate from ordinary-user browser actions.
"""
from __future__ import annotations

import argparse
import json
import os
import subprocess
import sys
from contextlib import contextmanager
from urllib.parse import quote, urlsplit

from openpyxl import load_workbook
from school_bootstrap_browser import ROOT, django_runtime, require
from school_structure_browser import click_document

OUT = ROOT / "tests/artifacts/school-staff-import-browser"
ROSTER = "/hr/staff/"
API = "/api/hr/v1/staff/import"
STAFF_NO = "IMPORT-0001"
NAME = "首次建档教职工甲"
DOCUMENT = "000000000000000009"  # synthetic fixture; never a real teacher


def write(name, value):
    OUT.mkdir(parents=True, exist_ok=True)
    (OUT / name).write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def seed():
    django_runtime()
    from django.contrib.auth.models import Permission
    from base.models import CompanyGroupAssignment
    from hr_staff.models import HrImportJob, HrPerson, HrStaffMaster

    parent = ROOT / "tests/artifacts/school-structure-browser"
    require(json.loads((parent / "mysql-seal.json").read_text())["status"] == "PASS", "Real structure seal required")
    require(not HrImportJob.objects.exists() and not HrPerson.objects.exists() and not HrStaffMaster.objects.exists(),
            "Import lane must start without personnel or staging jobs")
    data = json.loads((parent / "seed.json").read_text())
    for role in ("admin_a", "admin_b"):
        group = CompanyGroupAssignment.objects.get(user_id=data["roles"][role]["user_id"]).group
        permissions = Permission.objects.filter(codename__in=("hr.staff.view", "hr.staff.import"))
        require(permissions.count() == 2, "Canonical HR03 permissions missing")
        group.permissions.add(*permissions)
    group = CompanyGroupAssignment.objects.get(user_id=data["roles"]["viewer_a"]["user_id"]).group
    group.permissions.add(Permission.objects.get(codename="hr.staff.view"))
    write("seed.json", data)


def preview_seal():
    """Separate read-only process proves preview did not create a person."""
    django_runtime()
    from hr_staff.models import HrImportJob, HrPerson, HrStaffAssignment, HrStaffMaster
    data = json.loads((OUT / "receipt.json").read_text())
    job = HrImportJob.objects.get(pk=data["jobId"])
    require(job.valid_rows == 1 and job.failed_rows == 1 and job.status == "READY_TO_COMMIT", "Preview mismatch")
    require(not any(model.objects.exists() for model in (HrPerson, HrStaffMaster, HrStaffAssignment)), "Preview wrote authority facts")
    require(DOCUMENT not in json.dumps(list(job.rows.values_list("data_json", flat=True))), "Plaintext identity leaked into staging")
    write("preview-seal.json", {"status": "PASS", "personCount": 0, "validRows": 1, "failedRows": 1})


def browser_proof():
    from playwright.sync_api import sync_playwright

    base = os.environ.get("SCHOOL_BOOTSTRAP_BASE_URL", "http://127.0.0.1:8000").rstrip("/")
    require(urlsplit(base).hostname in ("localhost", "127.0.0.1"), "Loopback test runtime required")
    data = json.loads((OUT / "seed.json").read_text())
    evidence, failure = [], None

    def record(role, assertion, status):
        evidence.append({"role": role, "assertion": assertion, "http_status": status})
        write("evidence.json", evidence)

    def api(page, path, *, post=False):
        return page.evaluate("""async ({path, post}) => {
          const csrf = document.cookie.split(';').map(s=>s.trim()).find(s=>s.startsWith('csrftoken='));
          const response = await fetch(path, {credentials:'same-origin', method:post?'POST':'GET',
            headers:{'X-Requested-With':'XMLHttpRequest', 'X-CSRFToken':decodeURIComponent(csrf?.slice(10)||'')}});
          const text=await response.text();let payload=null;try{payload=JSON.parse(text)}catch(_){}
          return {status:response.status, payload, text};
        }""", {"path": path, "post": post})

    @contextmanager
    def login(browser, role):
        context = browser.new_context(viewport={"width": 1440, "height": 1000}, accept_downloads=True)
        context.tracing.start(screenshots=True, snapshots=True, sources=True)
        page, errors = context.new_page(), []
        page.on("pageerror", lambda error: errors.append(str(error)))
        try:
            page.goto(base + "/login/?next=" + quote(ROSTER, safe="/"))
            page.locator("#username").fill(data["roles"][role]["username"])
            page.locator("#password").fill(os.environ["SCHOOL_BOOTSTRAP_PASSWORD"] + role)
            click_document(page, page.locator("button.yk-login-submit"), base + ROSTER)
            require(any(cookie["name"] == "sessionid" for cookie in context.cookies()), "No session")
            yield page
            require(not errors, f"{role}: JavaScript errors {errors}")
        except BaseException:
            page.screenshot(path=str(OUT / f"failure-{role}.png"), full_page=True)
            raise
        finally:
            context.tracing.stop(path=str(OUT / f"trace-{role}.zip"))
            context.close()

    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            try:
                role = "admin_a"
                with login(browser, role) as page:
                    page.locator("#importToggle").click()
                    with page.expect_download() as downloaded:
                        page.locator("#downloadTemplate").click()
                    path = OUT / "template.xlsx"
                    downloaded.value.save_as(str(path))
                    book = load_workbook(path)
                    require("人员导入" in book.sheetnames and "本校组织岗位" in book.sheetnames, "Incomplete server template")
                    refs = list(book["本校组织岗位"].iter_rows(min_row=2, values_only=True))
                    reference = next((r for r in refs if r[0] == "TEACHING-OFFICE" and r[2] == "EDU-ADMIN-001"), None)
                    require(reference is not None, "Template did not include newly created formal placement")
                    require(all("学校 B" not in str(row) for row in refs), "Foreign school reference leaked")
                    record(role, "server-xlsx-template-with-own-hr02", 200)
                    fields = [cell.value for cell in book["人员导入"][1]]
                    good = {"staff_no": STAFF_NO, "legal_name": NAME, "gender_code": "U",
                            "staff_category_code": "ADMIN", "relationship_type": "REGULAR_EMPLOYMENT",
                            "document_number": DOCUMENT, "effective_from": reference[4],
                            "organization_code": reference[0], "position_code": reference[2], "fte": "1.00"}
                    bad = {**good, "staff_no": "IMPORT-ERROR", "legal_name": "", "document_number": ""}
                    for row_number, values in enumerate((good, bad), 2):
                        for column, key in enumerate(fields, 1):
                            cell = book["人员导入"].cell(row_number, column, values.get(key, "")); cell.data_type = "s"
                    upload_path = OUT / "staff-upload.xlsx"
                    book.save(upload_path); book.close()
                    page.locator("#importFile").set_input_files(str(upload_path))
                    with page.expect_response(lambda response: response.url == base + API and response.request.method == "POST") as uploaded:
                        page.locator("#importValidate").click()
                    require(uploaded.value.status == 201, f"Upload HTTP {uploaded.value.status}: {uploaded.value.text()}")
                    preview = uploaded.value.json()["data"]
                    require((preview["validRows"], preview["failedRows"]) == (1, 1), f"Preview mismatch: {preview}")
                    job_id = preview["jobId"]
                    write("receipt.json", {"jobId": job_id})
                    subprocess.run([sys.executable, str(ROOT / "scripts/school_staff_import_browser.py"), "preview-seal"], check=True)
                    page.wait_for_function("document.getElementById('importConfirmed').disabled === false")
                    require(page.locator("#importCommit").is_disabled(), "Commit did not require user confirmation")
                    require("第 3 行" in page.locator("#importIssues").inner_text(), "Original workbook row number missing")
                    record(role, "preview-staging-only-with-explicit-confirmation", 201)
                    with page.expect_download() as downloaded:
                        page.locator("#downloadImportErrors").click()
                    error_path = OUT / "errors.xlsx"; downloaded.value.save_as(str(error_path))
                    errors_book = load_workbook(error_path)
                    cells = list(errors_book.active.values)
                    require(cells[1][0] == "3", "Error workbook lost source row number")
                    require(DOCUMENT not in str(cells) and NAME not in str(cells), "Error workbook leaked personal fields")
                    errors_book.close()
                    record(role, "error-xlsx-download-no-personal-fields", 200)
                    page.locator("#importConfirmed").check()
                    with page.expect_response(lambda response: response.url == base + API + f"/{job_id}/commit" and response.request.method == "POST") as committed:
                        page.locator("#importCommit").click()
                    require(committed.value.status == 200, f"Commit failed {committed.value.text()}")
                    result = committed.value.json()["data"]
                    require(result["committedRows"] == 1 and result["status"] == "PARTIAL_FAILED", f"Commit mismatch {result}")
                    require(result["resultRows"][0]["staffId"], "No committed authority ID")
                    record(role, "confirmed-row-commit-to-formal-staff", 200)
                    row = page.locator("#rows tr").filter(has_text=STAFF_NO)
                    row.wait_for(state="visible")
                    require(all(text in row.inner_text() for text in (NAME, "教务处", "EDU-ADMIN-001", "在职")), "Roster placement readback mismatch")
                    page.reload(wait_until="domcontentloaded")
                    row = page.locator("#rows tr").filter(has_text=STAFF_NO); row.wait_for(state="visible")
                    page.wait_for_function("document.getElementById('importSummary').innerText.includes('部分行未写入')")
                    record(role, "reload-resumes-task-and-formal-roster", 200)
                    retried = api(page, API + f"/{job_id}/commit", post=True)
                    require(retried["status"] == 200 and retried["payload"]["data"]["committedRows"] == 1, "Duplicate commit was not idempotent")
                    record(role, "duplicate-commit-does-not-duplicate-staff", 200)
                    page.screenshot(path=str(OUT / "import-result-desktop.png"), full_page=True)
                    page.set_viewport_size({"width": 390, "height": 844})
                    page.locator("#importSummary").scroll_into_view_if_needed()
                    page.screenshot(path=str(OUT / "import-result-mobile.png"), full_page=True)
                    page.goto(base + "/hr/structure/positions")
                    post = page.locator("#hr-position-table tbody tr").filter(has_text="EDU-ADMIN-001")
                    post.wait_for(state="visible")
                    require("教务处" in post.inner_text() and "空缺" not in post.inner_text(), "HR02 did not reflect HR03 occupancy")
                    record(role, "hr02-position-reads-real-occupancy", 200)
                    page.screenshot(path=str(OUT / "occupied-position.png"), full_page=True)

                role = "viewer_a"
                with login(browser, role) as page:
                    row = page.locator("#rows tr").filter(has_text=STAFF_NO); row.wait_for(state="visible")
                    require(NAME in row.inner_text() and "教务处" in row.inner_text(), "Viewer cannot read final staff")
                    require(page.locator("#importToggle").is_disabled(), "Viewer import control enabled")
                    denied = api(page, API + f"/{job_id}/commit", post=True)
                    require(denied["status"] == 403, "Viewer commit not denied")
                    record(role, "formal-readback-with-import-write-denied", 403)

                role = "admin_b"
                with login(browser, role) as page:
                    own = api(page, "/api/hr/v1/staff?page=1")
                    require(own["status"] == 200 and own["payload"]["total"] == 0, "Foreign school roster leaked")
                    for suffix, post in (("", False), ("/errors", False), ("/commit", True)):
                        denied = api(page, API + f"/{job_id}" + suffix, post=post)
                        require(denied["status"] == 404, f"Foreign task access not concealed: {suffix}")
                    record(role, "foreign-task-and-personnel-concealed", 404)
            finally:
                browser.close()
    except BaseException as exc:
        failure = exc
    write("diagnostics.json", {"failure": repr(failure) if failure else None, "completedAssertions": len(evidence)})
    if failure:
        raise failure


def seal():
    django_runtime()
    from django.contrib.auth import get_user_model
    from employee.models import Employee
    from hr_staff.models import (
        HrEmploymentRelationship, HrImportJob, HrPerson, HrPersonIdentityDocument,
        HrStaffAssignment, HrStaffAuditEvent, HrStaffMaster,
    )
    from hr_structure.models import HrSchoolStructureInitialization

    seed = json.loads((OUT / "seed.json").read_text())
    job_id = json.loads((OUT / "receipt.json").read_text())["jobId"]
    evidence = json.loads((OUT / "evidence.json").read_text())
    expected = {
        ("admin_a", "server-xlsx-template-with-own-hr02"): 200,
        ("admin_a", "preview-staging-only-with-explicit-confirmation"): 201,
        ("admin_a", "error-xlsx-download-no-personal-fields"): 200,
        ("admin_a", "confirmed-row-commit-to-formal-staff"): 200,
        ("admin_a", "reload-resumes-task-and-formal-roster"): 200,
        ("admin_a", "duplicate-commit-does-not-duplicate-staff"): 200,
        ("admin_a", "hr02-position-reads-real-occupancy"): 200,
        ("viewer_a", "formal-readback-with-import-write-denied"): 403,
        ("admin_b", "foreign-task-and-personnel-concealed"): 404,
    }
    require({(r["role"], r["assertion"]): r["http_status"] for r in evidence} == expected and len(evidence) == len(expected),
            "Incomplete or duplicate browser evidence")
    tenant = seed["school_a"]
    job = HrImportJob.objects.get(tenant_id=tenant, pk=job_id)
    require(job.status == "PARTIAL_FAILED" and job.failed_rows == 1 and job.rows.filter(commit_status="COMMITTED").count() == 1,
            "Import job did not record actual row results")
    receipt = HrSchoolStructureInitialization.objects.get(tenant_id=tenant)
    staff = HrStaffMaster.objects.get(tenant_id=tenant, staff_no=STAFF_NO)
    require(staff.person_id.legal_name == NAME and staff.current_employment_status == "ACTIVE", "Wrong final personnel fact")
    relationship = HrEmploymentRelationship.objects.get(tenant_id=tenant, staff_id=staff)
    assignment = HrStaffAssignment.objects.get(tenant_id=tenant, employment_relationship_id=relationship)
    require(assignment.organization_id_id == receipt.department_version.organization_id_id
            and assignment.position_id_id == receipt.position_id and assignment.post_catalog_id_id == receipt.catalog_version_id,
            "Formal HR02 association missing")
    require(assignment.legacy_department_id is None and staff.primary_assignment_id == assignment.pk, "Legacy or projection mismatch")
    require(assignment.fte == 1, "Wrong real occupancy")
    for model in (HrPerson, HrStaffMaster, HrEmploymentRelationship, HrStaffAssignment, HrPersonIdentityDocument):
        require(model.objects.filter(tenant_id=tenant).count() == 1, f"Partial/duplicate facts in {model.__name__}")
        require(not model.objects.filter(tenant_id=seed["school_b"]).exists(), "Foreign authority mutated")
    actions = ("StaffImportValidated", "PersonCreated", "StaffMasterCreated", "EmploymentRelationshipStarted",
               "AssignmentCreated", "StaffImportRowCommitted", "StaffImportCompleted", "StaffImportIssuesDownloaded")
    audits = HrStaffAuditEvent.objects.filter(tenant_id=tenant, action__in=actions)
    require(audits.count() == len(actions) and set(audits.values_list("actor_user_id", flat=True)) == {seed["roles"]["admin_a"]["user_id"]},
            "Audit actor/actions mismatch")
    require(DOCUMENT not in json.dumps(list(job.rows.values_list("data_json", flat=True))), "Staging identity in plaintext")
    require(not Employee.objects.exists() and not get_user_model().objects.filter(is_superuser=True).exists(), "Legacy fixture/superuser bypass")
    write("mysql-seal.json", {"status": "PASS", "browserAssertions": len(expected), "staffCount": 1, "assignmentCount": 1,
                              "failedRows": 1, "legacyEmployees": 0, "actorAudits": audits.count(),
                              "productHead": os.environ["PRODUCT_HEAD_SHA"],
                              "testedCommit": subprocess.check_output(["git", "rev-parse", "HEAD"], text=True).strip()})
    print("HR03 XLSX import and formal HR02 placement MySQL seal PASS")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("phase", choices=("seed", "preview-seal", "browser", "seal"))
    {"seed": seed, "preview-seal": preview_seal, "browser": browser_proof, "seal": seal}[parser.parse_args().phase]()
