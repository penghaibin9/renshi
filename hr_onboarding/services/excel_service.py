"""
hr_onboarding/services/excel_service.py

Excel 导入（总册 §42 · 最小生产级实现）

流程：模板下载 → 上传 → staging 解析 → 业务校验 → error workbook → 确认 → async 执行 → result
禁止 Excel 直接绕过 Activation Service。

V1 支持：报到人员名单批量创建 case（source=HR04_HIRE 或 LEGACY_MIGRATION）。
"""

from __future__ import annotations

import io
import json
import logging
import uuid
from datetime import date, datetime
from typing import Optional

from django.core.files.base import ContentFile
from django.utils.translation import gettext as _
from django.utils import timezone

logger = logging.getLogger(__name__)

BATCH_CASE_TEMPLATE_COLS = [
    "legal_name",       # 姓名
    "source_type",      # 来源类型（HR04_HIRE / LEGACY_MIGRATION）
    "source_id",        # 来源标识
    "expected_report_date",  # 预计报到日
    "employment_type",  # 用工类型
    "staff_category",   # 人员类别
    "note",             # 备注
]

REQUIRED_COLS = {"legal_name", "source_type", "expected_report_date"}


class ExcelValidationError(Exception):
    def __init__(self, row: int, field: str, message: str):
        self.row = row
        self.field = field
        self.message = message
        super().__init__(f"row={row} field={field}: {message}")


class ExcelImportJob:
    """Excel 导入作业（staging → validate → confirm）。"""

    class Status:
        UPLOADED = "UPLOADED"
        VALIDATING = "VALIDATING"
        VALIDATION_FAILED = "VALIDATION_FAILED"
        READY_TO_COMMIT = "READY_TO_COMMIT"
        COMMITTING = "COMMITTING"
        COMPLETED = "COMPLETED"
        FAILED = "FAILED"

    def __init__(self, *, tenant_id: int, uploaded_by: int):
        self.tenant_id = tenant_id
        self.uploaded_by = uploaded_by
        self.job_id = str(uuid.uuid4())
        self.status = self.Status.UPLOADED
        self.rows = []
        self.errors = []
        self.result = {}

    def parse(
        self,
        uploaded_file,
        *,
        sheet_name: str = "Sheet1",
        expected_cols: Optional[list] = None,
    ) -> int:
        """解析上传文件为 staging rows（读第一 sheet，第一行作表头）。返回行数。"""
        try:
            import openpyxl
        except ImportError:
            raise ImportError(_("缺少 openpyxl 库。pip install openpyxl"))

        wb = openpyxl.load_workbook(uploaded_file, read_only=True)
        if sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[0]

        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [str(h or "").strip() if h else "" for h in next(rows_iter)]
        except StopIteration:
            return 0

        col_index = {}
        for i, h in enumerate(header):
            col_index[h] = i

        cols = expected_cols or BATCH_CASE_TEMPLATE_COLS
        self.status = self.Status.VALIDATING
        row_count = 0
        for row_data in rows_iter:
            row_no = row_count + 2  # 1-indexed, skip header
            record = {}
            for col in cols:
                idx = col_index.get(col)
                record[col] = str(row_data[idx]).strip() if idx is not None and row_data[idx] is not None else ""
            record["_row"] = row_no
            self.rows.append(record)
            row_count += 1
        wb.close()
        return row_count

    def validate(self) -> bool:
        """校验 staging 数据。返回 True 表示全部合法。"""
        from hr_onboarding.constants import CaseSourceType, EmploymentType, StaffCategoryCode

        self.errors = []
        for i, record in enumerate(self.rows):
            try:
                self._validate_row(i, record)
            except ExcelValidationError as exc:
                self.errors.append({"row": exc.row, "field": exc.field, "message": exc.message})

        if self.errors:
            self.status = self.Status.VALIDATION_FAILED
            return False
        self.status = self.Status.READY_TO_COMMIT
        return True

    def _validate_row(self, i, record):
        row = record["_row"]
        for col in REQUIRED_COLS:
            if not record.get(col):
                raise ExcelValidationError(row, col, "必填")

        st = record.get("source_type", "")
        if st and st not in CaseSourceType.values:
            raise ExcelValidationError(row, "source_type", f"无效来源: {st}")

        et = record.get("employment_type", "")
        if et and et not in EmploymentType.values:
            raise ExcelValidationError(row, "employment_type", f"无效用工类型: {et}")

        sc = record.get("staff_category", "")
        if sc and sc not in StaffCategoryCode.values:
            raise ExcelValidationError(row, "staff_category", f"无效人员类别: {sc}")

        dr = record.get("expected_report_date", "")
        if dr:
            try:
                date.fromisoformat(dr)
            except ValueError:
                raise ExcelValidationError(row, "expected_report_date", f"日期格式非法: {dr}")

    def error_workbook(self) -> bytes:
        """生成错误工作簿（xlsx）。"""
        try:
            import openpyxl
        except ImportError:
            return b""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "errors"
        ws.append(["row", "field", "message"])
        for e in self.errors:
            ws.append([e["row"], e["field"], e["message"]])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def commit_async(self) -> dict:
        """
        异步确认执行（逐行幂等：source_type+source_id unique 兜底）。
        禁止绕过 Activation Service（不直接建 Employment/Assignment）。
        """
        from hr_onboarding.services.case_service import CaseService

        self.status = self.Status.COMMITTING
        created = 0
        skipped = 0
        errors = []
        for record in self.rows:
            try:
                request = {
                    "source_type": record.get("source_type", "LEGACY_MIGRATION"),
                    "source_id": record.get("source_id", f"excel-{record['_row']}"),
                    "legal_name": record.get("legal_name", ""),
                    "employment_type": record.get("employment_type", "FULL_TIME"),
                    "staff_category": record.get("staff_category", "TEACHER"),
                    "expected_report_date": record.get("expected_report_date"),
                }
                service = CaseService(tenant_id=self.tenant_id)
                key = f"excel-import-{self.job_id}-{record['_row']}"
                result = service.create_case_from_handoff(request, idempotency_key=key)
                if result.get("created"):
                    created += 1
                else:
                    skipped += 1
            except Exception as exc:  # noqa: BLE001
                errors.append({"row": record["_row"], "error": str(exc)})
        if errors:
            self.status = self.Status.FAILED
        else:
            self.status = self.Status.COMPLETED
        self.result = {"created": created, "skipped": skipped, "errors": errors}
        return self.result

    def template_bytes(self) -> bytes:
        """生成空白导入模板（xlsx）。"""
        try:
            import openpyxl
        except ImportError:
            return b""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "import"
        ws.append(BATCH_CASE_TEMPLATE_COLS)
        # 示例行
        ws.append(["张三", "HR04_HIRE", "PH-2026-001", "2026-09-01", "FULL_TIME", "TEACHER", "示例"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
