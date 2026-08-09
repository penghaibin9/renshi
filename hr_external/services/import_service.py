"""
hr_external/services/import_service.py —— Excel 导入 staging + commit 执行（S3/S10，总册 §110）。

流程：template → upload → staging → validate → error workbook → preview → confirm
→ async execute → result ledger → audit（§110）。

- CSV（标准库）与 XLSX（openpyxl）原生解析，同一 HrExternalImportJob/Row 账本；
- confirm_job 只把 job 置 COMMITTING 并返回 202（00 §32：同步完成后伪装任务严格禁止）；
  真正执行由 job runner 调用 execute_commit()（分批事务 + checkpoint + 精确失败行，HR03 §24.3）。
- 逐行结果写入 HrExternalImportRow.status（COMMITTED/FAILED/SKIPPED）+ job error_summary_json 账本。
- 禁止 Excel 直接建账号/权限（§110/§24.4）：只允许建 Profile（身份根复用 HR03）。
"""

from __future__ import annotations

import csv
import io
from typing import Callable, Optional

from django.db import transaction

from hr_external.constants import (
    ExternalImportJobStatus,
    ExternalImportJobType,
    ExternalImportRowStatus,
)
from hr_external.integrations.hr03 import PersonProvider
from hr_external.models import HrExternalImportJob, HrExternalImportRow
from hr_external.services.profile_service import ProfileService

COMMIT_BATCH_SIZE = 100  # 分批事务大小（HR03 §24.3：checkpoint + 精确失败行）


class ImportValidationError(Exception):
    code = "INVALID_REQUEST"


class ImportCommitError(Exception):
    code = "IMPORT_COMMIT_FAILED"


class ImportService:
    @transaction.atomic
    def create_job(
        self,
        *,
        tenant_id: int,
        job_type: str,
        file_name: str = "",
        file_ref: str = "",
        template_version: str = "",
        created_by: Optional[int] = None,
    ) -> HrExternalImportJob:
        if job_type not in {t.value for t in ExternalImportJobType}:
            raise ImportValidationError("未知导入类型")
        return HrExternalImportJob.objects.create(
            tenant_id=tenant_id,
            job_type=job_type,
            file_name=file_name,
            file_ref=file_ref,
            template_version=template_version,
            created_by=created_by,
            status=ExternalImportJobStatus.UPLOADED,
        )

    @transaction.atomic
    def parse_csv_to_rows(self, job: HrExternalImportJob, content: bytes) -> int:
        """解析 CSV 到 staging rows（CSV 原生；XLSX 见 parse_spreadsheet 占位）。"""
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ImportValidationError("CSV 必须为 UTF-8 编码") from exc

        reader = csv.DictReader(io.StringIO(text))
        rows = []
        for idx, raw in enumerate(reader, start=2):  # 第 1 行为表头
            rows.append(
                HrExternalImportRow(
                    tenant_id=job.tenant_id,
                    job_id=job,
                    row_no=idx,
                    raw_json={k: v for k, v in raw.items() if v not in (None, "")},
                )
            )
        if rows:
            HrExternalImportRow.objects.bulk_create(rows)
        job.total_rows = len(rows)
        job.save(update_fields=["total_rows", "updated_at"])
        return len(rows)

    @transaction.atomic
    def parse_spreadsheet_to_rows(self, job: HrExternalImportJob, content: bytes) -> int:
        """XLSX 解析：openpyxl 读取首个 sheet，首行表头，写入与 CSV 相同的 staging 账本（§110/§33）。

        与 CSV 同链路：同一 HrExternalImportJob/Row 账本 + 逐行校验 + execute_commit。
        坏文件（非 XLSX/损坏）→ ImportValidationError（不静默）。
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.utils.exceptions import InvalidFileException
        except ImportError as exc:  # pragma: no cover —— 依赖已确认存在
            raise ImportValidationError("openpyxl 依赖不可用") from exc

        try:
            wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        except (InvalidFileException, ValueError, OSError) as exc:
            raise ImportValidationError(f"无法解析 XLSX 文件：{type(exc).__name__}") from exc

        sheet = wb.active
        if sheet is None or sheet.max_row < 1:
            raise ImportValidationError("XLSX 工作表为空")

        rows_iter = sheet.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            raise ImportValidationError("XLSX 缺少表头行")
        headers = [str(h).strip() if h is not None else "" for h in header]

        new_rows = []
        for row_no, values in enumerate(rows_iter, start=2):
            raw = {}
            for key, value in zip(headers, values):
                if value is None:
                    continue
                # 日期/数值转字符串；布尔转 "true"/"false"
                if hasattr(value, "strftime"):
                    raw[key] = value.strftime("%Y-%m-%d")
                elif isinstance(value, bool):
                    raw[key] = "true" if value else "false"
                else:
                    raw[key] = str(value)
            if not any(raw.values()):
                continue  # 空行跳过
            new_rows.append(
                HrExternalImportRow(
                    tenant_id=job.tenant_id,
                    job_id=job,
                    row_no=row_no,
                    raw_json=raw,
                )
            )
        if new_rows:
            HrExternalImportRow.objects.bulk_create(new_rows)
        job.total_rows = len(new_rows)
        job.save(update_fields=["total_rows", "updated_at"])
        return len(new_rows)

    @transaction.atomic
    def validate_job(
        self,
        job: HrExternalImportJob,
        validator: Callable[[dict], list],
    ) -> HrExternalImportJob:
        """逐行校验（validator(row_raw) -> list[str] issues）。"""
        job.status = ExternalImportJobStatus.VALIDATING
        job.save(update_fields=["status", "updated_at"])

        valid = invalid = 0
        issues_summary: dict[str, int] = {}
        for row in job.rows.all():
            issues = validator(row.raw_json) or []
            row.validation_issues = issues
            if issues:
                row.status = ExternalImportRowStatus.INVALID
                invalid += 1
                for issue in issues[:10]:
                    key = issue.split(":")[0][:40]
                    issues_summary[key] = issues_summary.get(key, 0) + 1
            else:
                row.status = ExternalImportRowStatus.VALID
                valid += 1
            row.save(update_fields=["validation_issues", "status"])

        job.success_count = valid
        job.failed_count = invalid
        job.error_summary_json = issues_summary
        job.status = (
            ExternalImportJobStatus.READY_TO_COMMIT
            if valid and invalid == 0
            else ExternalImportJobStatus.VALIDATION_FAILED
        )
        job.save(
            update_fields=[
                "status",
                "success_count",
                "failed_count",
                "error_summary_json",
                "updated_at",
            ]
        )
        return job

    def confirm_job(self, job: HrExternalImportJob) -> HrExternalImportJob:
        """confirm 只把 job 置 COMMITTING 并返回（§110/00 §32）。
        真正执行由 job runner 调用 execute_commit()；不得在此伪装完成。"""
        job.status = ExternalImportJobStatus.COMMITTING
        job.save(update_fields=["status", "updated_at"])
        return job

    def execute_commit(self, job: HrExternalImportJob) -> HrExternalImportJob:
        """异步执行 commit（PROFILE 类型）：分批事务 + 逐行结果账本（HR03 §24.3）。

        禁止：Excel 直接建账号/开放权限（§110）；本实现只建 Profile（身份根复用 HR03）。
        """
        if job.status != ExternalImportJobStatus.COMMITTING:
            raise ImportCommitError("job not in COMMITTING state")

        rows = list(job.rows.filter(status=ExternalImportRowStatus.VALID))
        person_provider = PersonProvider()
        profile_service = ProfileService()
        error_summary: dict[str, int] = {}
        committed = 0
        # 保留 validate 阶段已标记 INVALID 的失败行（§110 精确失败行账本）
        failed = job.rows.filter(status=ExternalImportRowStatus.INVALID).count()

        # 分批事务：每批原子提交，跨批 checkpoint（HR03 §24.3）
        # 计数只在整批事务成功提交后累计，保证回滚时账本精确。
        for start in range(0, len(rows), COMMIT_BATCH_SIZE):
            batch = rows[start : start + COMMIT_BATCH_SIZE]
            try:
                with transaction.atomic():
                    batch_committed, batch_failed, batch_errors = self._commit_batch(
                        job, batch, person_provider, profile_service
                    )
                committed += batch_committed
                failed += batch_failed
                for key, count in batch_errors.items():
                    error_summary[key] = error_summary.get(key, 0) + count
            except Exception:  # noqa: BLE001 —— 分批原子回滚，跨批 checkpoint 继续
                # 该批整体失败（行状态随事务回滚为 VALID，未累计计数）
                failed += len(batch)
                error_summary["BATCH_ATOMIC_FAILED"] = (
                    error_summary.get("BATCH_ATOMIC_FAILED", 0) + len(batch)
                )

        job.success_count = committed
        job.failed_count = failed
        job.error_summary_json = error_summary
        if failed == 0:
            job.status = ExternalImportJobStatus.COMPLETED
        else:
            job.status = ExternalImportJobStatus.PARTIAL_FAILED
        job.save(
            update_fields=[
                "status",
                "success_count",
                "failed_count",
                "error_summary_json",
                "updated_at",
            ]
        )
        return job

    @transaction.atomic
    def _commit_batch(
        self,
        job: HrExternalImportJob,
        batch: list,
        person_provider: PersonProvider,
        profile_service: ProfileService,
    ) -> tuple[int, int, dict]:
        """单批逐行 commit。返回 (committed, failed, error_summary)。整批原子。"""
        committed = failed = 0
        errors: dict[str, int] = {}
        for row in batch:
            ok, issue = self._commit_profile_row(
                job, row, person_provider, profile_service
            )
            if ok:
                committed += 1
            else:
                failed += 1
                row.status = ExternalImportRowStatus.FAILED
                row.validation_issues = [issue]
                row.save(update_fields=["status", "validation_issues"])
                key = (issue or "UNKNOWN")[:40]
                errors[key] = errors.get(key, 0) + 1
        return committed, failed, errors

    def _commit_profile_row(
        self,
        job: HrExternalImportJob,
        row: HrExternalImportRow,
        person_provider: PersonProvider,
        profile_service: ProfileService,
    ) -> tuple[bool, str]:
        """单行 PROFILE commit：先建/复用 HR03 Person，再建 HR08 Profile（幂等去重）。"""
        raw = row.raw_json
        legal_name = (raw.get("legalName") or "").strip()
        if not legal_name:
            row.status = ExternalImportRowStatus.FAILED
            row.validation_issues = ["legalName:必填"]
            row.save(update_fields=["status", "validation_issues"])
            return False, "legalName:必填"

        try:
            person_result = person_provider.create_person(
                tenant_id=job.tenant_id,
                legal_name=legal_name,
                preferred_name=raw.get("preferredName") or "",
                document_number=raw.get("documentNumber") or None,
                document_type=raw.get("documentType") or "NATIONAL_ID",
            )
            if not person_result.is_available:
                raise ImportCommitError(person_result.error_message)
            person_id = person_result.data["personId"]

            profile = profile_service.create_profile(
                tenant_id=job.tenant_id,
                person_id=person_id,
                primary_category_code=raw.get("primaryCategoryCode") or None,
                source_organization_name=raw.get("sourceOrganizationName") or "",
                source_position_title=raw.get("sourcePositionTitle") or "",
                industry_domain=raw.get("industryDomain") or "",
            )
            row.status = ExternalImportRowStatus.COMMITTED
            # 生产级（A21）：证件号明文只用于交 HR03 加密存储，HR08 staging 不保留明文。
            # raw_json 仅保留非敏感字段 + profileId 引用。
            safe_raw = {k: v for k, v in raw.items() if k not in ("documentNumber", "documentType")}
            row.raw_json = {**safe_raw, "_profileId": str(profile.id)}
            row.save(update_fields=["status", "raw_json"])
            return True, ""
        except Exception as exc:  # noqa: BLE001 —— 精确失败行
            return False, f"{getattr(exc, 'code', 'IMPORT_ROW_FAILED')}:{str(exc)[:80]}"
