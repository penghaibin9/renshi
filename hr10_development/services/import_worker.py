"""Controlled, idempotent Excel/legacy import parser for HR10."""

from __future__ import annotations

import hashlib
import json
import logging
from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO

from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)

MAX_IMPORT_ROWS = 10_000
SUPPORTED_TEMPLATE_VERSION = "V1"

TEMPLATE_SCHEMAS = {
    "EXCEL_PLAN": {
        "required": ("plan_no", "plan_type", "start_date", "end_date"),
        "date_fields": ("start_date", "end_date"),
        "integer_fields": ("owner_org_id", "staff_master_id"),
        "business_key": "plan_no",
        "target_model": "HrDevelopmentPlan",
    },
    "EXCEL_PROGRAM": {
        "required": ("program_code", "title", "activity_type"),
        "date_fields": (),
        "integer_fields": ("owner_org_id", "provider_org_id"),
        "business_key": "program_code",
        "target_model": "HrLearningProgram",
    },
    "EXCEL_PRACTICE": {
        "required": ("project_no", "title", "provider_org_id"),
        "date_fields": ("planned_start_date", "planned_end_date"),
        "integer_fields": ("provider_org_id", "owner_org_id", "capacity"),
        "business_key": "project_no",
        "target_model": "HrEnterprisePracticeProject",
    },
}

HEADER_ALIASES = {
    "计划编号": "plan_no",
    "计划类型": "plan_type",
    "开始日期": "start_date",
    "结束日期": "end_date",
    "归属组织ID": "owner_org_id",
    "教职工ID": "staff_master_id",
    "项目编码": "program_code",
    "项目编号": "project_no",
    "项目标题": "title",
    "活动类型": "activity_type",
    "机构ID": "provider_org_id",
    "计划开始日期": "planned_start_date",
    "计划结束日期": "planned_end_date",
    "容量": "capacity",
}


def run_import_job(job_id: int):
    """Parse one job under a row lock so duplicate worker delivery is harmless."""

    from hr10_development.legacy.import_job import HrDevelopmentImportJob

    try:
        with transaction.atomic():
            job = HrDevelopmentImportJob.objects.select_for_update().get(id=job_id)
            if job.status == "SUCCESS":
                return job
            job.status = "PARSE"
            job.started_at = job.started_at or timezone.now()
            job.retry_count += 1
            job.save(update_fields=["status", "started_at", "retry_count", "updated_at"])

        if "LEGACY_EMPLOYEE" in job.job_type:
            _parse_legacy_employee(job)
            final_status = "SUCCESS"
        elif job.job_type in TEMPLATE_SCHEMAS:
            _parse_excel(job)
            final_status = "PREVIEW"
        else:
            raise ValueError(f"UNSUPPORTED_JOB_TYPE: {job.job_type}")

        job.status = final_status
        job.completed_at = timezone.now() if final_status == "SUCCESS" else None
        job.save(update_fields=["status", "completed_at", "updated_at"])
        return job
    except HrDevelopmentImportJob.DoesNotExist:
        logger.error("Import job %s not found", job_id)
        return None
    except Exception as exc:
        logger.exception("Import job %s failed", job_id)
        HrDevelopmentImportJob.objects.filter(id=job_id).update(
            status="FAILED",
            result_summary_json={"errorCode": "IMPORT_PARSE_FAILED", "error": str(exc)[:2000]},
            completed_at=timezone.now(),
        )
        return None


def _parse_legacy_employee(job):
    """Stage legacy Employee qualification values without silently claiming zero."""

    from employee.models import Employee
    from hr10_development.legacy.staging import HrDevelopmentStagingRow

    employees = Employee.objects.filter(tenant_id=job.tenant_id, is_active=True)[:5000]
    created = 0
    for emp in employees:
        if not emp.qualification:
            continue
        HrDevelopmentStagingRow.objects.update_or_create(
            tenant_id=job.tenant_id,
            import_job_id=job.id,
            source_object_id=str(emp.id),
            defaults={
                "source_system": "LEGACY_EMPLOYEE",
                "source_table": "Employee",
                "source_field": "qualification",
                "raw_text": emp.qualification,
                "migration_trust_level": "UNKNOWN",
                "verification_status": "PENDING",
            },
        )
        created += 1

    job.total_rows = created
    job.processed_rows = created
    job.error_rows = 0
    job.result_summary_json = {
        "stagedRows": created,
        "sourceStatus": "AVAILABLE",
        "message": "Legacy Employee data staged for review",
    }
    job.save(
        update_fields=["total_rows", "processed_rows", "error_rows", "result_summary_json", "updated_at"]
    )


def _normalise_header(value) -> str:
    header = str(value or "").strip()
    return HEADER_ALIASES.get(header, header.lower().replace(" ", "_"))


def _json_value(value):
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _coerce_row(raw: dict, schema: dict) -> tuple[dict, list[str]]:
    parsed = {key: _json_value(value) for key, value in raw.items() if value not in (None, "")}
    errors = []

    for field in schema["required"]:
        if parsed.get(field) in (None, ""):
            errors.append(f"{field}: REQUIRED")

    for field in schema["date_fields"]:
        value = parsed.get(field)
        if value in (None, ""):
            continue
        try:
            parsed[field] = date.fromisoformat(str(value)[:10]).isoformat()
        except ValueError:
            errors.append(f"{field}: INVALID_DATE")

    for field in schema["integer_fields"]:
        value = parsed.get(field)
        if value in (None, ""):
            continue
        try:
            parsed[field] = int(value)
        except (TypeError, ValueError):
            errors.append(f"{field}: INVALID_INTEGER")

    start = parsed.get("start_date") or parsed.get("planned_start_date")
    end = parsed.get("end_date") or parsed.get("planned_end_date")
    if start and end and start > end:
        errors.append("date_range: START_AFTER_END")
    if isinstance(parsed.get("capacity"), int) and parsed["capacity"] < 0:
        errors.append("capacity: NEGATIVE")
    return parsed, errors


def _sha256(field_file) -> str:
    digest = hashlib.sha256()
    field_file.open("rb")
    try:
        for chunk in iter(lambda: field_file.read(1024 * 1024), b""):
            digest.update(chunk)
    finally:
        field_file.close()
    return digest.hexdigest()


def _save_error_workbook(job, headers: list[str], errors: list[dict]) -> str:
    if not errors:
        if job.error_workbook_path and default_storage.exists(job.error_workbook_path):
            default_storage.delete(job.error_workbook_path)
        return ""

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("errors")
    sheet.append(["row_number", *headers, "errors"])
    for item in errors:
        sheet.append(
            [item["rowNumber"], *[item["raw"].get(header) for header in headers], "; ".join(item["errors"])]
        )
    payload = BytesIO()
    workbook.save(payload)
    path = f"hr10/imports/{job.tenant_id}/errors/job-{job.id}.xlsx"
    if default_storage.exists(path):
        default_storage.delete(path)
    return default_storage.save(path, ContentFile(payload.getvalue()))


def _parse_excel(job):
    """Read a V1 XLSX workbook, validate every row and create replay-safe staging rows."""

    from hr10_development.legacy.staging import HrDevelopmentStagingRow

    if job.template_version != SUPPORTED_TEMPLATE_VERSION:
        raise ValueError(f"UNSUPPORTED_TEMPLATE_VERSION: {job.template_version}")
    if not job.source_file:
        raise ValueError("SOURCE_FILE_MISSING")
    if _sha256(job.source_file) != job.file_hash:
        raise ValueError("SOURCE_FILE_HASH_MISMATCH")

    job.source_file.open("rb")
    try:
        workbook = load_workbook(job.source_file, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        raw_headers = next(rows, None)
        if not raw_headers:
            raise ValueError("EMPTY_WORKBOOK")
        headers = [_normalise_header(value) for value in raw_headers]
        if any(not header for header in headers) or len(headers) != len(set(headers)):
            raise ValueError("INVALID_OR_DUPLICATE_HEADERS")

        schema = TEMPLATE_SCHEMAS[job.job_type]
        missing_headers = sorted(set(schema["required"]) - set(headers))
        if missing_headers:
            raise ValueError(f"MISSING_HEADERS: {','.join(missing_headers)}")

        valid_rows = []
        error_rows = []
        seen_keys = set()
        total = 0
        for row_number, values in enumerate(rows, start=2):
            if not any(value not in (None, "") for value in values):
                continue
            total += 1
            if total > MAX_IMPORT_ROWS:
                raise ValueError(f"ROW_LIMIT_EXCEEDED: {MAX_IMPORT_ROWS}")
            raw = {headers[index]: _json_value(value) for index, value in enumerate(values) if index < len(headers)}
            parsed, errors = _coerce_row(raw, schema)
            business_key = str(parsed.get(schema["business_key"], "")).strip()
            if business_key and business_key in seen_keys:
                errors.append(f"{schema['business_key']}: DUPLICATE_IN_WORKBOOK")
            seen_keys.add(business_key)
            if errors:
                error_rows.append({"rowNumber": row_number, "raw": raw, "errors": errors})
            else:
                valid_rows.append((row_number, raw, parsed))
    finally:
        job.source_file.close()

    with transaction.atomic():
        HrDevelopmentStagingRow.objects.filter(
            tenant_id=job.tenant_id,
            import_job_id=job.id,
            source_system="EXCEL",
        ).delete()
        HrDevelopmentStagingRow.objects.bulk_create(
            [
                HrDevelopmentStagingRow(
                    tenant_id=job.tenant_id,
                    source_system="EXCEL",
                    source_table=job.job_type,
                    source_field="ROW",
                    source_object_id=f"{job.template_version}:{row_number}",
                    raw_text=json.dumps(raw, ensure_ascii=False, sort_keys=True),
                    parsed_data=parsed,
                    migration_trust_level="DECLARED",
                    target_model=schema["target_model"],
                    import_job_id=job.id,
                    verification_status="PENDING",
                )
                for row_number, raw, parsed in valid_rows
            ],
            batch_size=500,
        )

    error_path = _save_error_workbook(job, headers, error_rows)
    job.total_rows = total
    job.processed_rows = len(valid_rows)
    job.error_rows = len(error_rows)
    job.warning_rows = 0
    job.checkpoint_row = total + 1 if total else 1
    job.error_workbook_path = error_path
    job.result_summary_json = {
        "templateVersion": job.template_version,
        "sourceHash": job.file_hash,
        "totalRows": total,
        "validRows": len(valid_rows),
        "errorRows": len(error_rows),
        "previewReady": not error_rows,
        "replaySafe": True,
    }
    job.save(
        update_fields=[
            "total_rows",
            "processed_rows",
            "error_rows",
            "warning_rows",
            "checkpoint_row",
            "error_workbook_path",
            "result_summary_json",
            "updated_at",
        ]
    )
