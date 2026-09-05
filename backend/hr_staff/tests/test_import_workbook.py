"""Transport regression tests also run without Django; no fabricated DB evidence."""
import io
import re
import unittest
import zipfile
from datetime import datetime
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from hr_staff.services import import_workbook as transport


class ImportWorkbookTests(unittest.TestCase):
    def workbook(self, row=None, headers=None):
        book = Workbook()
        sheet = book.active
        sheet.title = transport.DATA_SHEET
        sheet.append(headers or list(transport.COLUMNS))
        if row is not None:
            sheet.append(row)
        out = io.BytesIO()
        book.save(out)
        return out.getvalue()

    def row(self):
        return ["000012", "教师甲", "F", "1990-01-02", "000000000000000001", "TEACHER",
                "REGULAR_EMPLOYMENT", "2026-09-01", "TEACHING", "TEACHER-001", "1.00"]

    def repack(self, raw, add=None, replace=None):
        out = io.BytesIO()
        with zipfile.ZipFile(io.BytesIO(raw)) as original, zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as updated:
            for name in original.namelist():
                updated.writestr(name, (replace or {}).get(name, original.read(name)))
            for name, data in (add or {}).items():
                updated.writestr(name, data)
        return out.getvalue()

    def test_text_identifiers_keep_leading_zeroes(self):
        row = transport.parse_upload(self.workbook(self.row()), "staff.xlsx")[0]
        self.assertEqual(row["staff_no"], "000012")
        self.assertEqual(row["document_number"], "000000000000000001")

    def test_excel_date_becomes_iso_date(self):
        row = self.row(); row[3] = datetime(1990, 1, 2)
        self.assertEqual(transport.parse_upload(self.workbook(row), "staff.xlsx")[0]["birth_date"], "1990-01-02")

    def test_numeric_identifiers_are_rejected_before_precision_loss(self):
        for column in (0, 4, 8, 9):
            with self.subTest(column=column):
                row = self.row(); row[column] = 123456789012345678
                with self.assertRaisesRegex(transport.ImportFileError, "文本"):
                    transport.parse_upload(self.workbook(row), "staff.xlsx")

    def test_formulas_and_excel_errors_are_not_accepted_as_values(self):
        for value in ('=1+1', '#DIV/0!'):
            row = self.row(); row[1] = value
            with self.subTest(value=value), self.assertRaisesRegex(transport.ImportFileError, "公式|Excel"):
                transport.parse_upload(self.workbook(row), "staff.xlsx")

    def test_whitespace_headers_and_bom_csv_are_normalized(self):
        result = transport.parse_upload("\ufeff staff_no , legal_name \n0001, 教师甲 \n".encode(), "staff.csv")
        self.assertEqual(result, [{"staff_no": "0001", "legal_name": "教师甲"}])

    def test_duplicate_unknown_and_empty_headers_fail_closed(self):
        for headers in (["legal_name", "legal_name"], ["legal_name", "tenant_id"], ["legal_name", ""]):
            with self.subTest(headers=headers), self.assertRaises(transport.ImportFileError):
                transport.parse_upload((",".join(headers) + "\nA,B\n").encode(), "file.csv")

    def test_extra_csv_cells_and_invalid_utf8_are_rejected(self):
        for raw in (b"legal_name\nA,B\n", b"\xff\xfe"):
            with self.assertRaises(transport.ImportFileError):
                transport.parse_upload(raw, "file.csv")

    def test_empty_file_and_empty_data_sheet_are_rejected(self):
        for name, raw in (("f.csv", b""), ("f.xlsx", self.workbook())):
            with self.subTest(name=name), self.assertRaises(transport.ImportFileError):
                transport.parse_upload(raw, name)

    def test_blank_middle_row_preserves_source_row_numbers(self):
        raw = b"staff_no,legal_name\nA,A\n,\nC,C\n\n"
        result = transport.parse_upload(raw, "f.csv")
        self.assertEqual(len(result), 3)
        self.assertEqual(result[1], {"staff_no": "", "legal_name": ""})

    def test_named_sheet_required_no_silent_wrong_sheet_import(self):
        raw = self.repack(self.workbook(self.row()), replace={
            "xl/workbook.xml": self._zip_read(self.workbook(self.row()), "xl/workbook.xml").replace("人员导入".encode(), b"wrong")})
        with self.assertRaises(transport.ImportFileError):
            transport.parse_upload(raw, "f.xlsx")

    @staticmethod
    def _zip_read(raw, name):
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            return archive.read(name)

    def test_active_content_and_external_links_rejected(self):
        for name in ("xl/vbaProject.bin", "xl/embeddings/object.bin", "xl/externalLinks/externalLink1.xml"):
            with self.subTest(name=name), self.assertRaises(transport.ImportFileError):
                transport.parse_upload(self.repack(self.workbook(self.row()), add={name: b"forbidden"}), "f.xlsx")
        raw = self.repack(self.workbook(self.row()), add={"xl/worksheets/_rels/sheet1.xml.rels": b'<Relationships><Relationship TargetMode="External" Target="https://example.invalid"/></Relationships>'})
        with self.assertRaisesRegex(transport.ImportFileError, "外部"):
            transport.parse_upload(raw, "f.xlsx")

    def test_archive_traversal_and_entity_declarations_rejected(self):
        for name, content in (("../bad.xml", b"bad"), ("xl/_rels/evil.rels", b'<!DOCTYPE a [<!ENTITY x "boom">]><a>&x;</a>')):
            with self.assertRaises(transport.ImportFileError):
                transport.parse_upload(self.repack(self.workbook(self.row()), add={name: content}), "f.xlsx")

    def test_compressed_and_uncompressed_size_limits(self):
        with patch.object(transport, "MAX_IMPORT_BYTES", 8), self.assertRaises(transport.ImportFileError):
            transport.parse_upload(b"longer-than-limit", "f.csv")
        with patch.object(transport, "MAX_UNPACKED_BYTES", 32), self.assertRaises(transport.ImportFileError):
            transport.parse_upload(self.workbook(self.row()), "f.xlsx")

    def test_limits_include_row_count_and_cell_length(self):
        with patch.object(transport, "MAX_IMPORT_ROWS", 1), self.assertRaises(transport.ImportFileError):
            transport.parse_upload(b"legal_name\nA\nB\n", "f.csv")
        row = self.row(); row[1] = "x" * 513
        with self.assertRaises(transport.ImportFileError):
            transport.parse_upload(self.workbook(row), "f.xlsx")

    def test_disguised_zip_and_unsupported_extensions(self):
        for name in ("f.xlsx", "f.xls", "f.xlsm", "f.html"):
            with self.subTest(name=name), self.assertRaises(transport.ImportFileError):
                transport.parse_upload(b"not-an-xlsx", name)

    def test_template_has_no_example_people_and_uses_only_canonical_fields(self):
        raw = transport.template_workbook([("OFFICE", "教务处", "POST-1", "教务岗", "2026-09-06")])
        book = load_workbook(io.BytesIO(raw))
        sheet = book[transport.DATA_SHEET]
        self.assertEqual(tuple(cell.value for cell in sheet[1]), transport.COLUMNS)
        self.assertNotIn("legacy_department_id", transport.COLUMNS)
        self.assertIsNone(sheet["A2"].value)
        self.assertEqual(sheet["A2"].number_format, "@")
        self.assertEqual(book["本校组织岗位"]["B2"].value, "教务处")
        self.assertEqual(sheet.freeze_panes, "A2")
        book.close()

    def test_error_report_is_text_and_not_reimportable(self):
        raw = transport.error_workbook([(2, "position_code", "VALIDATION_ERROR", '=HYPERLINK("x")')])
        book = load_workbook(io.BytesIO(raw), data_only=False)
        self.assertEqual(book.active["D2"].data_type, "s")
        self.assertNotIn("document_number", str(list(book.active.values)))
        book.close()
        with self.assertRaises(transport.ImportFileError):
            transport.parse_upload(raw, "errors.xlsx")

    def _rewrite_sheet(self, raw, transform, name="xl/worksheets/sheet1.xml"):
        return self.repack(raw, replace={name: transform(self._zip_read(raw, name))})

    def test_actual_row_beyond_declared_dimension_is_rejected(self):
        raw = self.workbook(self.row())
        extra = b'<row r="6000"><c r="A6000" t="inlineStr"><is><t>not silently ignored</t></is></c></row>'
        raw = self._rewrite_sheet(raw, lambda xml: xml.replace(b"</sheetData>", extra + b"</sheetData>"))
        with self.assertRaisesRegex(transport.ImportFileError, "5000|行号"):
            transport.parse_upload(raw, "staff.xlsx")

    def test_actual_column_beyond_declared_dimension_is_rejected(self):
        raw = self.workbook(self.row())
        extra = b'<row r="3"><c r="Z3" t="inlineStr"><is><t>not silently ignored</t></is></c></row>'
        raw = self._rewrite_sheet(raw, lambda xml: xml.replace(b"</sheetData>", extra + b"</sheetData>"))
        with self.assertRaisesRegex(transport.ImportFileError, "列"):
            transport.parse_upload(raw, "staff.xlsx")

    def test_understated_dimension_does_not_hide_valid_data(self):
        raw = self._rewrite_sheet(self.workbook(self.row()),
                                  lambda xml: re.sub(rb'<dimension ref="[^"]+"', b'<dimension ref="A1:A1"', xml))
        result = transport.parse_upload(raw, "staff.xlsx")
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["staff_no"], "000012")
        self.assertEqual(result[0]["fte"], "1.00")

    def test_duplicate_row_coordinates_are_rejected(self):
        def duplicate(xml):
            row = re.search(rb'<row r="2".*?</row>', xml).group(0)
            return xml.replace(b"</sheetData>", row + b"</sheetData>")
        raw = self._rewrite_sheet(self.workbook(self.row()), duplicate)
        with self.assertRaisesRegex(transport.ImportFileError, "行号"):
            transport.parse_upload(raw, "staff.xlsx")

    def test_duplicate_cell_coordinates_are_rejected(self):
        def duplicate(xml):
            row = re.search(rb'<row r="2".*?</row>', xml).group(0)
            cell = re.search(rb'<c r="A2".*?</c>', row).group(0)
            return xml.replace(row, row.replace(b"</row>", cell + b"</row>"))
        raw = self._rewrite_sheet(self.workbook(self.row()), duplicate)
        with self.assertRaisesRegex(transport.ImportFileError, "列"):
            transport.parse_upload(raw, "staff.xlsx")

    def test_named_sheet_not_first_is_validated_by_its_relationship(self):
        book = Workbook()
        book.active.title = "填写说明"
        book.active["A1"] = "instructions"
        sheet = book.create_sheet(transport.DATA_SHEET)
        sheet.append(list(transport.COLUMNS)); sheet.append(self.row())
        out = io.BytesIO(); book.save(out); book.close()
        raw = out.getvalue()
        self.assertEqual(transport.parse_upload(raw, "staff.xlsx")[0]["staff_no"], "000012")
        bad = self._rewrite_sheet(raw, lambda xml: xml.replace(b'r="2"', b'r="6000"'),
                                  name="xl/worksheets/sheet2.xml")
        with self.assertRaises(transport.ImportFileError):
            transport.parse_upload(bad, "staff.xlsx")

    def test_mismatched_cell_row_is_not_silently_accepted(self):
        raw = self._rewrite_sheet(self.workbook(self.row()), lambda xml: xml.replace(b'r="B2"', b'r="B3"'))
        with self.assertRaisesRegex(transport.ImportFileError, "错位"):
            transport.parse_upload(raw, "staff.xlsx")

    def test_bad_worksheet_target_is_rejected(self):
        raw = self.workbook(self.row())
        rels = self._zip_read(raw, "xl/_rels/workbook.xml.rels")
        bad = self.repack(raw, replace={"xl/_rels/workbook.xml.rels":
                          rels.replace(b"/xl/worksheets/sheet1.xml", b"/foreign/sheet1.xml")})
        with self.assertRaises(transport.ImportFileError):
            transport.parse_upload(bad, "staff.xlsx")
