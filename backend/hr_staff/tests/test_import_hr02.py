"""Real MySQL contracts for the HR02 -> HR03 verified-staff import boundary."""
import io
import json
from concurrent.futures import ThreadPoolExecutor
from datetime import timedelta
from decimal import Decimal
from threading import Barrier
from types import SimpleNamespace
from unittest.mock import patch

from django.contrib.auth.models import Permission
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import close_old_connections, connections
from django.test import TestCase, TransactionTestCase, override_settings
from django.utils import timezone
from django.urls import resolve, reverse
from openpyxl import load_workbook

from employee.models import Employee
from hr_staff.constants import ImportJobStatus
from hr_staff.models import (
    HrEmploymentRelationship, HrPerson, HrPersonIdentityDocument,
    HrStaffAssignment, HrStaffAuditEvent, HrStaffMaster,
)
from hr_staff.services.import_service import ImportService, ImportStateConflict, StaffMasterRowApplier
from hr_staff.services.import_validation import StaffImportValidator, basic_errors, peak_usage
from hr_staff.services.import_workbook import COLUMNS, DATA_SHEET, template_workbook
from hr_structure.models import HrPosition, HrPositionReservation
from hr_structure.tests.test_initial_setup import Fixture, SETTINGS


class ImportFixture(Fixture):
    def setUp(self):
        super().setUp()
        self.receipt, _ = self.command()
        for user in (self.admin, self.other_admin):
            user.company_group_assignments.get().group.permissions.add(*Permission.objects.filter(
                codename__in=("hr.staff.import", "hr.staff.view")))
        self.viewer.company_group_assignments.get().group.permissions.add(
            Permission.objects.get(codename="hr.staff.view"))

    def row(self, number="IMPORT-001", **changes):
        return {"staff_no": number, "legal_name": "导入教师" + number,
                "organization_code": "OFFICE", "position_code": "ACADEMIC-001",
                "effective_from": self.today.isoformat(), "fte": "1.00", **changes}

    def ready(self, rows):
        service = ImportService(self.school.pk, actor_user_id=self.admin.pk)
        job = service.create_job(template_key="staff_master_hr02")
        job.checkpoint = {"upload_actor_user_id": self.admin.pk}
        job.save(update_fields=["checkpoint"])
        validator = StaffImportValidator(self.school.pk, rows)
        service.parse_rows(job, rows)
        service.validate_rows(job, validator, row_enricher=validator.enrich)
        return service, job

    def commit(self, service, job):
        return service.commit(job, StaffMasterRowApplier(self.school.pk, actor_user_id=self.admin.pk))

    @staticmethod
    def xlsx(rows):
        book = load_workbook(io.BytesIO(template_workbook()))
        sheet = book[DATA_SHEET]
        for row_no, row in enumerate(rows, 2):
            for col, name in enumerate(COLUMNS, 1):
                cell = sheet.cell(row_no, col, row.get(name, ""))
                cell.data_type = "s"
        out = io.BytesIO(); book.save(out); book.close()
        return out.getvalue()

    def upload(self, browser, rows):
        return browser.post("/api/v1/hr/staff/import", {"file": SimpleUploadedFile("staff.xlsx", self.xlsx(rows))},
                            HTTP_X_CSRFTOKEN=browser.cookies["csrftoken"].value)


@override_settings(**SETTINGS)
class FormalStaffImportTests(ImportFixture, TestCase):
    def test_preview_is_staging_only_and_identity_is_encrypted(self):
        document = "000000000000000001"
        _, job = self.ready([self.row(document_number=document)])
        self.assertEqual(job.valid_rows, 1)
        stage = job.rows.get()
        self.assertNotIn(document, json.dumps(stage.data_json))
        self.assertNotIn("document_number", stage.data_json)
        self.assertIn("_document_ciphertext", stage.data_json)
        self.assertEqual(stage.data_json["_structure_snapshot"]["position"], self.receipt.position_id)
        self.assertFalse(HrPerson.objects.exists())
        self.assertFalse(HrStaffMaster.objects.exists())

    def test_commit_creates_linked_formal_facts_and_actor_audits(self):
        service, job = self.ready([self.row(document_number="000000000000000001")])
        self.assertEqual(self.commit(service, job), {"committed": 1, "failed": 0, "total": 1})
        staff = HrStaffMaster.objects.get(tenant_id=self.school.pk, staff_no="IMPORT-001")
        relation = HrEmploymentRelationship.objects.get(staff_id=staff)
        assignment = HrStaffAssignment.objects.get(employment_relationship_id=relation)
        self.assertEqual(assignment.organization_id_id, self.receipt.department_version.organization_id_id)
        self.assertEqual(assignment.position_id_id, self.receipt.position_id)
        self.assertEqual(assignment.post_catalog_id_id, self.receipt.catalog_version_id)
        self.assertIsNone(assignment.legacy_department_id)
        self.assertEqual(assignment.fte, Decimal("1"))
        self.assertEqual(staff.current_employment_status, "ACTIVE")
        self.assertEqual(staff.primary_assignment_id, assignment.pk)
        actions = ("PersonCreated", "StaffMasterCreated", "EmploymentRelationshipStarted", "AssignmentCreated", "StaffImportRowCommitted", "StaffImportCompleted")
        events = HrStaffAuditEvent.objects.filter(tenant_id=self.school.pk, action__in=actions)
        self.assertEqual(events.count(), 6)
        self.assertEqual(set(events.values_list("actor_user_id", flat=True)), {self.admin.pk})
        self.assertFalse(Employee.objects.exists())
        self.assertEqual(HrPersonIdentityDocument.objects.count(), 1)

    def test_same_file_and_database_duplicates_do_not_overwrite(self):
        service, job = self.ready([self.row(), self.row()])
        self.assertEqual((job.valid_rows, job.failed_rows), (1, 1))
        self.assertEqual(self.commit(service, job)["committed"], 1)
        _, another = self.ready([self.row()])
        self.assertEqual(another.status, ImportJobStatus.VALIDATION_FAILED)
        self.assertEqual(HrStaffMaster.objects.count(), 1)

    def test_likely_identity_duplicates_are_flagged_in_preview(self):
        row = self.row(legal_name="同名教师", birth_date="1990-01-01")
        _, job = self.ready([row, {**row, "staff_no": "DIFFERENT"}])
        self.assertEqual(job.failed_rows, 1)
        self.assertTrue(job.issues.filter(field_code="legal_name").exists())

    def test_legacy_department_is_not_used_to_bypass_formal_placement(self):
        _, job = self.ready([self.row(organization_code="", position_code="", legacy_department_id="7")])
        self.assertEqual(job.status, ImportJobStatus.VALIDATION_FAILED)
        self.assertFalse(HrStaffMaster.objects.exists())

    def test_unknown_cross_school_and_mismatched_codes_fail_in_preview(self):
        for changes in ({"organization_code": "FOREIGN-ORG"}, {"position_code": "FOREIGN-POST"},
                        {"organization_code": "SCH"}):
            with self.subTest(changes=changes):
                _, job = self.ready([self.row(**changes)])
                self.assertEqual(job.status, ImportJobStatus.VALIDATION_FAILED)
        self.assertFalse(HrPerson.objects.exists())

    def test_future_start_and_nonfinite_fte_rejected(self):
        for changes in ({"effective_from": (self.today + timedelta(days=1)).isoformat()},
                        {"fte": "NaN"}, {"fte": "Infinity"}, {"fte": "0"}, {"fte": "0.001"},
                        {"relationship_type": "EXTERNAL_PART_TIME"}):
            with self.subTest(changes=changes):
                _, job = self.ready([self.row(**changes)])
                self.assertEqual(job.status, ImportJobStatus.VALIDATION_FAILED)

    def test_birth_after_employment_start_is_rejected(self):
        _, job = self.ready([self.row(effective_from="2000-01-01", birth_date="2001-01-01")])
        self.assertEqual(job.status, ImportJobStatus.VALIDATION_FAILED)
        self.assertTrue(job.issues.filter(field_code="effective_from").exists())

    def test_business_day_is_explicit_not_process_timezone(self):
        tomorrow = self.today + timedelta(days=1)
        row = self.row(effective_from=tomorrow.isoformat())
        self.assertNotIn("effective_from", basic_errors(row, today=tomorrow))
        self.assertIn("effective_from", basic_errors(row, today=self.today))

    def test_identity_with_only_separators_is_rejected_before_staging(self):
        self.assertIn("document_number", basic_errors(self.row(document_number="------")))

    def test_preview_batches_staging_updates_not_one_update_per_row(self):
        from django.db import connection
        from django.test.utils import CaptureQueriesContext
        service = ImportService(self.school.pk, actor_user_id=self.admin.pk)
        job = service.create_job(template_key="validation-query-budget")
        service.parse_rows(job, [{"staff_no": str(i)} for i in range(501)])
        with CaptureQueriesContext(connection) as captured:
            service.validate_rows(job, lambda row: {"legal_name": "必填"})
        updates = [query["sql"] for query in captured if query["sql"].lstrip().upper().startswith("UPDATE")
                   and "hr_staff_hrimportrow" in query["sql"].lower()]
        self.assertEqual(len(updates), 3)
        self.assertEqual(job.failed_rows, 501)
        self.assertEqual(job.issues.count(), 501)

    def test_valid_hold_blocks_unreserved_import(self):
        HrPositionReservation.objects.create(tenant_id=self.school.pk, reservation_no="HELD-1",
            position_id=self.receipt.position, source_domain="HR04", source_business_type="TEST",
            source_business_id="held", idempotency_key="held", expires_at=timezone.now() + timedelta(hours=1))
        _, job = self.ready([self.row()])
        self.assertEqual(job.status, ImportJobStatus.VALIDATION_FAILED)
        self.assertTrue(job.issues.filter(field_code="position_code").exists())

    def test_preview_accounts_for_capacity_across_the_file(self):
        _, job = self.ready([self.row("ONE"), self.row("TWO")])
        self.assertEqual((job.valid_rows, job.failed_rows), (1, 1))
        self.assertEqual(job.issues.get().row_no, 3)

    def test_post_preview_freeze_is_revalidated_without_partial_person(self):
        service, job = self.ready([self.row()])
        HrPosition.objects.filter(pk=self.receipt.position_id).update(lifecycle_status="FROZEN")
        result = self.commit(service, job)
        self.assertEqual(result, {"committed": 0, "failed": 1, "total": 1})
        self.assertFalse(HrPerson.objects.exists())
        self.assertTrue(job.issues.filter(field_code="position_code").exists())

    def test_post_preview_capacity_consumption_is_revalidated(self):
        service, job = self.ready([self.row("FIRST")])
        StaffMasterRowApplier(self.school.pk, actor_user_id=self.admin.pk)(self.row("COMPETITOR"), {})
        self.assertEqual(self.commit(service, job)["committed"], 0)
        self.assertEqual(HrStaffMaster.objects.count(), 1)
        self.assertEqual(HrPerson.objects.count(), 1)

    def test_late_audit_failure_rolls_back_whole_person_and_reports_safe_error(self):
        from hr_staff.services.audit_service import write_audit_event
        service, job = self.ready([self.row(document_number="000000000000000002")])
        def audited(**kwargs):
            if kwargs["action"] == "StaffImportRowCommitted":
                raise RuntimeError("private SQL/password diagnostics")
            return write_audit_event(**kwargs)
        with patch("hr_staff.services.audit_service.write_audit_event", side_effect=audited):
            self.assertEqual(self.commit(service, job)["committed"], 0)
        self.assertFalse(HrPerson.objects.exists())
        self.assertFalse(HrStaffMaster.objects.exists())
        self.assertFalse(HrStaffAssignment.objects.exists())
        self.assertNotIn("private SQL", job.issues.get().message)

    def test_encrypted_document_cannot_be_rebound_to_another_job(self):
        _, one = self.ready([self.row("ONE", document_number="000000000000000003")])
        service, two = self.ready([self.row("TWO", document_number="000000000000000004")])
        stage = two.rows.get()
        stage.data_json["_document_ciphertext"] = one.rows.get().data_json["_document_ciphertext"]
        stage.save(update_fields=["data_json"])
        self.assertEqual(self.commit(service, two)["committed"], 0)
        self.assertFalse(HrPerson.objects.exists())

    def test_finished_job_is_idempotent_and_old_executor_is_fenced(self):
        service, job = self.ready([self.row()])
        self.commit(service, job)
        self.assertEqual(self.commit(service, job)["committed"], 1)
        self.assertEqual(HrStaffMaster.objects.count(), 1)
        with self.assertRaises(ImportStateConflict):
            service._assert_executor(SimpleNamespace(status=ImportJobStatus.COMMITTING), {"commit_token": "new"}, "old")

    def test_bounded_commits_resume_same_job_without_false_completion(self):
        service = ImportService(self.school.pk, actor_user_id=self.admin.pk)
        job = service.create_job(template_key="chunk-contract")
        service.parse_rows(job, [{"staff_no": f"CHUNK-{i}"} for i in range(5)])
        service.validate_rows(job, lambda row: {})
        applied = []
        first = service.commit(job, lambda row, checkpoint: applied.append(row["staff_no"]), batch_size=2)
        self.assertEqual(first, {"committed": 2, "failed": 0, "total": 5, "pending": 3})
        job.refresh_from_db()
        self.assertEqual(job.status, ImportJobStatus.READY_TO_COMMIT)
        self.assertIsNone(job.committed_at)
        self.assertFalse(HrStaffAuditEvent.objects.filter(action="StaffImportCompleted").exists())
        service.commit(job, lambda row, checkpoint: applied.append(row["staff_no"]), batch_size=2)
        last = service.commit(job, lambda row, checkpoint: applied.append(row["staff_no"]), batch_size=2)
        self.assertEqual(last, {"committed": 5, "failed": 0, "total": 5})
        self.assertEqual(len(set(applied)), 5)
        self.assertEqual(HrStaffAuditEvent.objects.filter(action="StaffImportCompleted").count(), 1)

    def test_status_exposes_recovery_only_for_stale_lease_and_owner_can_resume(self):
        service, job = self.ready([self.row()])
        job.status = ImportJobStatus.COMMITTING
        job.checkpoint = {**job.checkpoint, "commit_token": "lost-process",
                          "commit_heartbeat_at": (timezone.now() - timedelta(hours=1)).isoformat()}
        job.save(update_fields=["status", "checkpoint"])
        browser = self.login(self.admin)
        path = "/api/v1/hr/staff/import/" + str(job.pk)
        before = browser.get(path)
        self.assertEqual(before.status_code, 200)
        self.assertTrue(before.json()["data"]["canResume"])
        self.assertNotIn("lost-process", before.content.decode())
        resumed = browser.post(path + "/commit", HTTP_X_CSRFTOKEN=browser.cookies["csrftoken"].value)
        self.assertEqual(resumed.status_code, 200, resumed.content)
        self.assertEqual(resumed.json()["data"]["committedRows"], 1)
        self.assertFalse(resumed.json()["data"]["canResume"])
        self.assertEqual(HrStaffMaster.objects.count(), 1)

    def test_final_audit_failure_can_resume_without_recreating_committed_person(self):
        from hr_staff.services.audit_service import write_audit_event
        service, job = self.ready([self.row()])
        def fail_final_only(**kwargs):
            if kwargs["action"] == "StaffImportCompleted":
                raise RuntimeError("completion audit unavailable")
            return write_audit_event(**kwargs)
        with patch("hr_staff.services.audit_service.write_audit_event", side_effect=fail_final_only):
            with self.assertRaises(RuntimeError):
                self.commit(service, job)
        job.refresh_from_db()
        self.assertEqual(job.status, ImportJobStatus.COMMITTING)
        self.assertEqual(job.rows.filter(commit_status="COMMITTED").count(), 1)
        self.assertEqual(HrStaffMaster.objects.count(), 1)
        self.assertFalse(HrStaffAuditEvent.objects.filter(action="StaffImportCompleted").exists())
        job.checkpoint = {**job.checkpoint,
                          "commit_heartbeat_at": (timezone.now() - timedelta(hours=1)).isoformat()}
        job.save(update_fields=["checkpoint"])
        browser = self.login(self.admin)
        path = "/api/v1/hr/staff/import/" + str(job.pk)
        observed = browser.get(path).json()["data"]
        self.assertTrue(observed["canResume"])
        self.assertEqual(observed["pendingRows"], 0)
        response = browser.post(path + "/commit", HTTP_X_CSRFTOKEN=browser.cookies["csrftoken"].value)
        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual(response.json()["data"]["status"], "COMPLETED")
        self.assertEqual(HrPerson.objects.count(), 1)
        self.assertEqual(HrStaffMaster.objects.count(), 1)
        self.assertEqual(HrStaffAssignment.objects.count(), 1)
        self.assertEqual(HrStaffAuditEvent.objects.filter(action="StaffImportCompleted").count(), 1)

    def test_active_executor_remains_busy_despite_status_read(self):
        service, job = self.ready([self.row()])
        job.status = ImportJobStatus.COMMITTING
        job.checkpoint = {**job.checkpoint, "commit_token": "active-process",
                          "commit_heartbeat_at": timezone.now().isoformat()}
        job.save(update_fields=["status", "checkpoint"])
        browser = self.login(self.admin)
        path = "/api/v1/hr/staff/import/" + str(job.pk)
        self.assertFalse(browser.get(path).json()["data"]["canResume"])
        denied = browser.post(path + "/commit", HTTP_X_CSRFTOKEN=browser.cookies["csrftoken"].value)
        self.assertEqual(denied.status_code, 409)
        self.assertFalse(HrPerson.objects.exists())

    def test_peak_usage_counts_interval_overlap_not_lifetime_sum(self):
        tomorrow = self.today + timedelta(days=1)
        self.assertEqual(peak_usage([(self.today, tomorrow, Decimal("1")), (tomorrow, None, Decimal("1"))], self.today), (1, Decimal("1")))
        self.assertEqual(peak_usage([(self.today, None, Decimal("1")), (tomorrow, None, Decimal("0.5"))], self.today), (2, Decimal("1.5")))

    def test_real_api_template_upload_commit_and_status_use_current_school(self):
        browser = self.login(self.admin)
        template = browser.get("/api/v1/hr/staff/import/template")
        self.assertEqual(template.status_code, 200)
        self.assertIn("spreadsheetml", template["Content-Type"])
        uploaded = self.upload(browser, [self.row(), self.row("BAD", legal_name="")])
        self.assertEqual(uploaded.status_code, 201, uploaded.content)
        data = uploaded.json()["data"]
        self.assertEqual((data["validRows"], data["failedRows"]), (1, 1))
        self.assertFalse(HrStaffMaster.objects.exists())
        path = "/api/v1/hr/staff/import/" + data["jobId"]
        committed = browser.post(path + "/commit", HTTP_X_CSRFTOKEN=browser.cookies["csrftoken"].value)
        self.assertEqual(committed.status_code, 200, committed.content)
        self.assertEqual(committed.json()["data"]["committed"], 1)
        status = browser.get(path).json()["data"]
        self.assertEqual(status["status"], "PARTIAL_FAILED")
        self.assertEqual(status["resultRows"][0]["staffId"], str(HrStaffMaster.objects.get().pk))
        errors = browser.get(path + "/errors")
        self.assertEqual(errors.status_code, 200)
        workbook = load_workbook(io.BytesIO(errors.content))
        self.assertEqual(workbook.active["A2"].value, "3")
        workbook.close()
        self.assertTrue(HrStaffAuditEvent.objects.filter(action="StaffImportIssuesDownloaded", actor_user_id=self.admin.pk).exists())

    def test_api_rejects_missing_csrf_and_foreign_jobs_and_viewer_writes(self):
        owner = self.login(self.admin)
        upload = self.upload(owner, [self.row()])
        path = "/api/v1/hr/staff/import/" + upload.json()["data"]["jobId"]
        self.assertEqual(owner.post(path + "/commit").status_code, 403)
        foreign = self.login(self.other_admin)
        for tail in ("", "/errors"):
            self.assertEqual(foreign.get(path + tail).status_code, 404)
        self.assertEqual(foreign.post(path + "/commit", HTTP_X_CSRFTOKEN=foreign.cookies["csrftoken"].value).status_code, 404)
        viewer = self.login(self.viewer)
        self.assertEqual(viewer.get("/api/v1/hr/staff/import/template").status_code, 403)
        self.assertEqual(viewer.post(path + "/commit", HTTP_X_CSRFTOKEN=viewer.cookies["csrftoken"].value).status_code, 403)
        self.assertFalse(HrPerson.objects.exists())

    def test_same_school_other_importer_cannot_read_or_commit_owner_task(self):
        owner = self.login(self.admin)
        response = self.upload(owner, [self.row()])
        path = "/api/v1/hr/staff/import/" + response.json()["data"]["jobId"]
        self.viewer.company_group_assignments.get().group.permissions.add(
            Permission.objects.get(codename="hr.staff.import"))
        other = self.login(self.viewer)
        for tail in ("", "/errors"):
            self.assertEqual(other.get(path + tail).status_code, 404)
        self.assertEqual(other.post(path + "/commit", HTTP_X_CSRFTOKEN=other.cookies["csrftoken"].value).status_code, 404)
        self.assertFalse(HrStaffMaster.objects.exists())

    def test_public_import_routes_are_canonical_and_legacy_is_redirect_only(self):
        from hr_staff.models import HrImportJob
        browser = self.login(self.admin)
        canonical = reverse("hr03-api-staff-import-template")
        self.assertEqual(canonical, "/api/v1/hr/staff/import/template")
        self.assertEqual(resolve(canonical).func.__module__, "hr_staff.api.imports")
        response = browser.get("/api/hr/v1/staff/import/template")
        self.assertEqual(response.status_code, 308)
        self.assertEqual(response["Location"], canonical)
        response = browser.post("/api/hr/v1/staff/import", {},
            HTTP_X_CSRFTOKEN=browser.cookies["csrftoken"].value)
        self.assertEqual(response.status_code, 308)
        self.assertEqual(response["Location"], "/api/v1/hr/staff/import")
        self.assertFalse(HrImportJob.objects.exists())
        self.assertFalse(HrPerson.objects.exists())

    def test_api_status_has_no_plaintext_document_or_birth_date(self):
        browser = self.login(self.admin)
        uploaded = self.upload(browser, [self.row(document_number="000000000000000005", birth_date="1990-01-01")])
        self.assertEqual(uploaded.status_code, 201)
        content = uploaded.content.decode()
        self.assertNotIn("000000000000000005", content)
        self.assertNotIn("1990-01-01", content)


@override_settings(**SETTINGS)
class ConcurrentFormalImportTests(ImportFixture, TransactionTestCase):
    def test_two_jobs_cannot_fill_one_position_twice(self):
        _, first = self.ready([self.row("RACE-ONE")])
        _, second = self.ready([self.row("RACE-TWO")])
        barrier = Barrier(2)
        school_id, actor_id = self.school.pk, self.admin.pk
        def run(job_id):
            close_old_connections()
            try:
                service = ImportService(school_id, actor_id)
                job = service.job_for_id(job_id)
                barrier.wait(timeout=15)
                return service.commit(job, StaffMasterRowApplier(school_id, actor_id))
            finally:
                connections.close_all()
        with ThreadPoolExecutor(max_workers=2) as pool:
            results = list(pool.map(run, [first.pk, second.pk]))
        self.assertEqual(sum(result["committed"] for result in results), 1)
        self.assertEqual(sum(result["failed"] for result in results), 1)
        self.assertEqual(HrStaffMaster.objects.filter(tenant_id=school_id).count(), 1)
        self.assertEqual(HrPerson.objects.filter(tenant_id=school_id).count(), 1)
        self.assertEqual(HrStaffAssignment.objects.filter(position_id=self.receipt.position).count(), 1)
