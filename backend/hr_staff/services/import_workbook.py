"""Bounded CSV/XLSX transport for HR03; no model writes or formula evaluation.

Workbook imports require the named data sheet. Identifiers stay text, formulas,
external relationships and active content are rejected before staging. Error
workbooks contain coordinates/messages, never uploaded identity documents.
"""
from __future__ import annotations

import csv
import io
import posixpath
import re
import zipfile
from datetime import date, datetime

from defusedxml import ElementTree
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 5000
MAX_UNPACKED_BYTES = 32 * 1024 * 1024
DATA_SHEET = "人员导入"
COLUMNS = (
    "staff_no", "legal_name", "gender_code", "birth_date", "document_number",
    "staff_category_code", "relationship_type", "effective_from",
    "organization_code", "position_code", "fte",
)
# Historical CSV remains parseable, but the web validator rejects unmapped old
# department IDs. Verified offline migrations retain their existing service.
LEGACY_COLUMNS = ("legacy_department_id",)
LABELS = (
    "工号（必填，文本）", "姓名（必填）", "性别 M/F/O/U", "出生日期 YYYY-MM-DD",
    "证件号（可选，文本）", "人员类别代码", "聘用关系代码", "任职生效日（必填）",
    "HR02 组织编码（必填）", "HR02 岗位编码（必填）", "任职工作量（默认 1.00）",
)
TEXT_IDENTIFIERS = {"staff_no", "document_number", "organization_code", "position_code"}


class ImportFileError(ValueError):
    """Message is safe for an upload response; never echo workbook values."""


def _headers(values):
    names = [str(value or "").strip() for value in values]
    if not names or any(not name for name in names):
        raise ImportFileError("表头缺失或含空列名，请使用下载的导入模板。")
    if len(names) != len(set(names)):
        raise ImportFileError("表头存在重复列名。")
    if set(names) - set(COLUMNS + LEGACY_COLUMNS):
        raise ImportFileError("文件含不支持的列，请使用下载的导入模板。")
    if "legal_name" not in names:
        raise ImportFileError("缺少 legal_name 姓名列。")
    return names


def _bounded_rows(rows):
    # Keep interior empty rows so source Excel/CSV row numbers remain exact.
    while rows and not any(rows[-1].values()):
        rows.pop()
    if not rows:
        raise ImportFileError("没有可导入的数据行。")
    return rows


SHEET_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
REL_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"


def _data_sheet_path(archive, names):
    """Resolve the named worksheet; never assume that it is sheet1.xml."""
    workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    matches = [node for node in workbook.iter(SHEET_NS + "sheet")
               if node.get("name") == DATA_SHEET]
    if len(matches) != 1:
        raise ImportFileError("缺少或重复的“人员导入”工作表，请使用下载的模板。")
    relation_id = matches[0].get(REL_NS + "id")
    relations = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    targets = [node for node in relations if node.get("Id") == relation_id]
    if len(targets) != 1 or targets[0].get("TargetMode", "").lower() == "external":
        raise ImportFileError("人员导入工作表的引用无效。")
    target = targets[0].get("Target", "")
    if not target or "\\" in target:
        raise ImportFileError("人员导入工作表的引用无效。")
    path = posixpath.normpath(target.lstrip("/") if target.startswith("/")
                             else posixpath.join("xl", target))
    if not path.startswith("xl/worksheets/") or path not in names:
        raise ImportFileError("人员导入工作表的引用无效。")
    return path


class _DataSheetBounds:
    """Validate actual coordinates before a bounded reader can hide excess data.

    Worksheet dimension metadata is not authoritative. Require monotonic row
    and cell coordinates so duplicate rows/cells cannot overwrite silently.
    Attribute-less rows/cells use the same sequential defaults as the reader.
    """

    def __init__(self):
        self.row = 0
        self.column = 0

    def start(self, item):
        if item.tag == SHEET_NS + "row":
            raw = item.get("r", str(self.row + 1))
            if not re.fullmatch(r"[1-9][0-9]{0,6}", raw):
                raise ImportFileError("工作表含无效行号。")
            row = int(raw)
            if row <= self.row or row > MAX_IMPORT_ROWS + 1:
                raise ImportFileError("数据表行号重复、顺序异常或超过 5000 个数据行。")
            self.row, self.column = row, 0
        elif item.tag == SHEET_NS + "c":
            coordinate = item.get("r")
            if coordinate is None:
                row, column = self.row, self.column + 1
            else:
                match = re.fullmatch(r"([A-Z]{1,3})([1-9][0-9]{0,6})", coordinate)
                if match is None:
                    raise ImportFileError("工作表含无效单元格坐标。")
                column = 0
                for letter in match[1]:
                    column = column * 26 + ord(letter) - ord("A") + 1
                row = int(match[2])
            if row != self.row or row < 1 or column <= self.column or column > len(COLUMNS):
                raise ImportFileError("数据表含重复、错位或表头范围之外的列。")
            self.column = column


def _check_archive(raw):
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            members = archive.infolist()
            names = [member.filename for member in members]
            if len(members) > 200 or len(names) != len(set(names)):
                raise ImportFileError("工作簿结构过于复杂或含重复文件。")
            if sum(member.file_size for member in members) > MAX_UNPACKED_BYTES:
                raise ImportFileError("工作簿解压后过大，请拆分导入。")
            if "xl/workbook.xml" not in names or "[Content_Types].xml" not in names:
                raise ImportFileError("上传内容不是有效的 xlsx 工作簿。")
            data_path = _data_sheet_path(archive, names)
            bounds = _DataSheetBounds()
            for member in members:
                name = member.filename
                lower = name.lower()
                if (name.startswith("/") or "\\" in name or posixpath.normpath(name).startswith("..")
                        or member.flag_bits & 1 or member.compress_type not in (0, 8)):
                    raise ImportFileError("工作簿包含不支持的压缩内容。")
                if any(part in lower for part in ("vba", "externallinks/", "embeddings/", "activex/", "connections", "querytables/")):
                    raise ImportFileError("不允许宏、外部数据源或嵌入对象。")
                if lower.endswith((".xml", ".rels")):
                    with archive.open(member) as stream:
                        stack = []
                        for event, item in ElementTree.iterparse(stream, events=("start", "end")):
                            if event == "start":
                                stack.append(item)
                                if lower.endswith(".rels") and item.get("TargetMode", "").lower() == "external":
                                    raise ImportFileError("工作簿含外部链接，请移除后重新上传。")
                                if name == data_path:
                                    bounds.start(item)
                            else:
                                stack.pop()
                                item.clear()
                                if stack:
                                    stack[-1].remove(item)
    except ImportFileError:
        raise
    except Exception as exc:
        raise ImportFileError("工作簿损坏、加密或结构不受支持。") from exc


def parse_upload(raw: bytes, filename: str) -> list[dict]:
    if len(raw) > MAX_IMPORT_BYTES:
        raise ImportFileError("上传文件不能超过 5 MB。")
    if filename.lower().endswith(".csv"):
        try:
            reader = csv.reader(io.StringIO(raw.decode("utf-8-sig")), strict=True)
            names = _headers(next(reader, []))
            rows = []
            for values in reader:
                if len(rows) >= MAX_IMPORT_ROWS:
                    raise ImportFileError("单次最多 5000 个数据行，请拆分导入。")
                if not values:
                    values = [""] * len(names)
                if len(values) != len(names):
                    raise ImportFileError(f"第 {len(rows) + 2} 行列数与表头不一致。")
                if any(len(value) > 512 for value in values):
                    raise ImportFileError(f"第 {len(rows) + 2} 行单元格内容过长。")
                rows.append(dict(zip(names, (value.strip() for value in values))))
            return _bounded_rows(rows)
        except (UnicodeError, csv.Error) as exc:
            raise ImportFileError("CSV 必须使用 UTF-8 编码且结构完整。") from exc
    if not filename.lower().endswith(".xlsx"):
        raise ImportFileError("仅支持 xlsx 或 UTF-8 CSV；不支持 xls、宏工作簿或其他格式。")
    _check_archive(raw)
    workbook = None
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=False, keep_links=False)
        if DATA_SHEET not in workbook.sheetnames:
            raise ImportFileError("缺少“人员导入”工作表，请使用下载的模板。")
        sheet = workbook[DATA_SHEET]
        if (sheet.max_row and sheet.max_row > MAX_IMPORT_ROWS + 1) or (sheet.max_column and sheet.max_column > len(COLUMNS)):
            raise ImportFileError("数据表超过 5000 行或含多余列，请删除多余格式与数据。")
        # Actual sheet coordinates were checked above; discard untrusted used-
        # range metadata so a forged small dimension cannot hide populated cells.
        sheet.reset_dimensions()
        iterator = sheet.iter_rows(max_row=MAX_IMPORT_ROWS + 2, max_col=len(COLUMNS) + 1)
        header = list(next(iterator, ()))
        while header and header[-1].value is None:
            header.pop()
        if any(cell.data_type in ("f", "e") for cell in header):
            raise ImportFileError("表头不能含公式或 Excel 错误。")
        names = _headers([cell.value for cell in header])
        rows = []
        for cells in iterator:
            row_no = len(rows) + 2
            if not any(cell.value is not None for cell in cells):
                rows.append(dict.fromkeys(names, ""))
                continue
            if len(rows) >= MAX_IMPORT_ROWS:
                raise ImportFileError("单次最多 5000 个数据行。")
            if any(cell.value is not None for cell in cells[len(names):]):
                raise ImportFileError(f"第 {row_no} 行含表头之外的数据。")
            values = []
            for name, cell in zip(names, cells):
                value = cell.value
                if cell.data_type in ("f", "e"):
                    raise ImportFileError(f"第 {row_no} 行含公式或 Excel 错误，请粘贴为值。")
                if value is not None and name in TEXT_IDENTIFIERS and not isinstance(value, str):
                    raise ImportFileError(f"第 {row_no} 行的 {name} 必须按文本填写，避免编号丢零或精度损失。")
                if isinstance(value, (datetime, date)):
                    if name not in ("birth_date", "effective_from"):
                        raise ImportFileError(f"第 {row_no} 行包含放错列的日期。")
                    value = value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
                value = "" if value is None else str(value).strip()
                if len(value) > 512:
                    raise ImportFileError(f"第 {row_no} 行单元格内容过长。")
                values.append(value)
            rows.append(dict(zip(names, values)))
        return _bounded_rows(rows)
    except ImportFileError:
        raise
    except Exception as exc:
        raise ImportFileError("无法读取工作簿，请重新下载模板并粘贴数据。") from exc
    finally:
        if workbook is not None:
            workbook.close()


def _text(cell, value):
    cell.value = str(value if value is not None else "")
    # Explicit text type also protects strings starting =, +, -, @.
    cell.data_type = "s"
    cell.number_format = "@"


def _style(sheet, widths):
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="245B90")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[1].height = 34
    for index, width in enumerate(widths, 1):
        from openpyxl.utils import get_column_letter
        sheet.column_dimensions[get_column_letter(index)].width = width


def template_workbook(reference_rows=()) -> bytes:
    book = Workbook()
    sheet = book.active
    sheet.title = DATA_SHEET
    for index, (code, label) in enumerate(zip(COLUMNS, LABELS), 1):
        _text(sheet.cell(1, index), code)
        sheet.cell(1, index).comment = Comment(label, "跃科人事")
        for row in range(2, 102):
            sheet.cell(row, index).number_format = "@"
    for column, values in ((3, "M,F/O/U".replace("/", ",")), (6, "TEACHER,ADMIN,ENGINEERING_TECHNICAL,EXPERIMENTAL,LIBRARY_ARCHIVES,LOGISTICS,OTHER")):
        validation = DataValidation(type="list", formula1='"' + values + '"', allow_blank=True)
        sheet.add_data_validation(validation)
        validation.add(f"{sheet.cell(1, column).column_letter}2:{sheet.cell(1, column).column_letter}5001")
    _style(sheet, [20, 20, 14, 18, 26, 24, 29, 20, 25, 25, 16])
    guide = book.create_sheet("填写说明")
    guide.append(["字段", "填写要求"])
    for code, label in zip(COLUMNS, LABELS):
        guide.append([code, label])
    guide.append(["用途", "仅导入已核验在职教职工。新招聘入职请走 HR05。不会创建登录账号。"])
    guide.append(["生效日期", "必填 YYYY-MM-DD，不晚于今天；组织岗位在该日期必须正式生效。"])
    guide.append(["错误修正", "先预检，再确认提交有效行。错误清单按原文件行号定位；修正失败行后另行上传，勿重复上传已成功行。"])
    guide.append(["参考范围", "本校组织岗位页最多提供 2000 个当前岗位，不代表可占用容量；更多岗位请在 HR02 台账核对。"])
    guide.append(["安全", "工号、组织岗位编码、证件号均为文本；不接受公式、外部链接或宏。模板不含真实人员信息。"])
    _style(guide, [28, 80])
    for row in guide.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        guide.row_dimensions[row[1].row].height = 42
    refs = book.create_sheet("本校组织岗位")
    refs.append(["organization_code", "组织名称", "position_code", "岗位目录名称", "核验日期"])
    for row in reference_rows:
        for index, value in enumerate(row, 1):
            _text(refs.cell(refs.max_row + 1 if index == 1 else refs.max_row, index), value)
    _style(refs, [25, 32, 25, 32, 20])
    out = io.BytesIO()
    book.save(out)
    return out.getvalue()


def error_workbook(issues) -> bytes:
    book = Workbook()
    sheet = book.active
    sheet.title = "错误清单"
    sheet.append(["原文件行号", "字段", "错误代码", "处理提示"])
    for issue in issues:
        index = sheet.max_row + 1
        for col, value in enumerate(issue, 1):
            _text(sheet.cell(index, col), value)
    _style(sheet, [18, 26, 35, 74])
    for row in sheet.iter_rows(min_row=2):
        row[3].alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row[0].row].height = 34
    out = io.BytesIO()
    book.save(out)
    return out.getvalue()
